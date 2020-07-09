from django.shortcuts import render
from .models import ModelNetworking
from django.http import HttpResponse
# from django.views.generic import View
from django.urls import resolve
import os
import json
from django.http import JsonResponse
# from rest_framework import generics, status
# from .serializers import ModelNetworkingSerial
# from rest_framework.response import Response
import csv
import platform
from openpyxl import Workbook
from uuid import uuid4
import re

HOSTNAME = platform.uname()[1].lower().strip()
PATH_CSV_NETWORK = '/home/henry/files/{}'

def IndexNetworking(request):
    uri_logo = uri_file('ico_network.jpg')
    uri_ico = uri_file('logo.ico')
    uri_ap = uri_api()
    return render(request, 'index.html', {'uri_logo': uri_logo, 'uri_ico': uri_ico, 'uri_ap':uri_ap})

def save_data_csv(data):
    id_ = str(uuid4()).split('-')[0]
    file = '{}.csv'.format(id_)

    with open(PATH_CSV_NETWORK.format(file), 'w', encoding='utf-8') as myfile:
        wr = csv.writer(myfile, quoting=csv.QUOTE_ALL)
        list_email = [dt[7] for dt in data]
        list_corr = []
        for eml in list_email:
            try:
                email_pattern = r'[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,4}'
                eml = re.findall(email_pattern, eml)
                list_corr.extend(eml)
            except Exception as e:
                print(e)

        list_corr = set(list_corr)
        list_corr = list(list_corr)
        for co in list_corr:
            wr.writerow([co])
        uri = uri_file(file)
        return uri


def save_data_xlsx(data):
    id_ = str(uuid4()).split('-')[0]
    file = '{}.xlsx'.format(id_)
    # with open(PATH_CSV_NETWORK.format(file), 'w', encoding='utf-8') as myfile:
    #     wr = csv.writer(myfile, quoting=csv.QUOTE_ALL)
    #     for dt in data:
    #         wr.writerow(dt)
    #     uri = uri_file(file)
    #     return uri
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'RAZON SOCIAL'
    ws['B1'] = 'REPRESENTANTES PRINCIPALES'
    ws['C1'] = 'PRESENCIA ONLINE'
    ws['D1'] = 'FECHA DE APERTURA'
    ws['E1'] = 'TIPO DE EMPRESA'
    ws['H1'] = 'CATEGORIA'
    ws['I1'] = 'TLF MOBILE'
    ws['J1'] = 'EMAIL'
    ws['K1'] = 'WEBSITE ESTIMADO'
    ws['L1'] = 'FUENTES RELACIONADAS'
    ws['M1'] = 'FACEBOOK FANPAGE'
    ws['N1'] = 'WEB SERVER'
    ws['O1'] = 'WEB FRAMEWORK'
    ws['P1'] = 'JS FRAMEWORK'
    ws['Q1'] = 'CMS'
    ws['R1'] = 'N° Empleados/Fecha'
    ws['S1'] = 'RUC'
    ws['T1'] = 'RUBRO'
    ws['U1'] = 'Direccion'
    ws['V1'] = 'LENGUAJE PROGRAMMING'
    ws['W1'] = 'STATUS COMPANY'
    ws['X1'] = 'TWITTER PAGE'

    for item in data:
        try:
            ws.append(item)
        except Exception as e:
            debug(e)
    wb.save(PATH_CSV_NETWORK.format(file))
    uri = uri_file(file)
    return uri

def uri_file(file_csv):
    url = f'/api/file/?name={file_csv}'
    return url

def uri_api():
    url = '/api/networking/'
    return url

def file_csv(request):
    if request.method == 'GET':
        file = request.GET['name']
        if '.xlsx' in file:
            with open(PATH_CSV_NETWORK.format(file), 'rb') as myfile:
                response = HttpResponse(myfile, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename={0}'.format(file)
                return response
        elif '.jpg' in file or '.ico' in file:
            with open(PATH_CSV_NETWORK.format(file), 'rb') as myfile:
                response = HttpResponse(myfile, content_type='image/jpg')
                return response
        elif '.csv' in file:
            with open(PATH_CSV_NETWORK.format(file), 'rb') as myfile:
                response = HttpResponse(myfile, content_type='text/csv')
                response['Content-Disposition'] = 'attachment; filename={0}'.format(file)
                return response


def Networking(request):
    if request.method == 'POST':
        data = request.body
        data = eval(data)
        # data = json.loads(data)
        prese_ =  data.get('prese_', '') # 3 opciones
        tipo_ =  data.get('tipo_', '') # opciones
        ctg_ =  data.get('ctg_', '') # opciones
        repre_ =  data.get('repre_', '') #filter vacio lleno
        report_ =  data.get('report_', '')# vacio, lleno
        linke_ =  data.get('linke_', '')# vacio lleno
        tlf_ =  data.get('tlf_', '') # vacio lleno
        fb_ =  data.get('fb_', '')# vacio lleno
        email_ =  data.get('email_', '')# vacio llen
        web_ =  data.get('web_', '')# vacio lleno
        tipo_tech =  data.get('tipo_tech', '')# vacio lleno

        if tipo_ == '0' or tipo_ == '2' or tipo_ == 'Todas':
            tipo_ = ''
        if prese_ == '0' or prese_ == '2' or prese_ == 'Todas':
            prese_ = ''
        if ctg_ == '0' or ctg_ == '2' or ctg_ == 'Todas':
            ctg_ = ''
        if tipo_tech == '0' or tipo_tech == '2' or tipo_tech == 'Todas':
            tipo_tech = ''
        if ctg_ == 'MODA / VESTIMENTA':
            ctg_ = ''
        queryset = ModelNetworking.objects.filter(
                    presence_online__icontains = prese_,
                    tamano_company__icontains = tipo_,
                    category__icontains = ctg_,
                    all_tech__icontains = tipo_tech,
            )

        if repre_ == '1':
            queryset = queryset.exclude(representantes='')
        elif repre_ == '2':
            queryset = queryset.filter(representantes='')

        if linke_ == '1':
            queryset = queryset.exclude(uri_reporte_linkedin='')
        elif linke_ == '2':
            queryset = queryset.filter(uri_reporte_linkedin='')

        if tlf_ == '1':
            queryset = queryset.exclude(tlf_contacto='')
        elif tlf_ == '2':
            queryset = queryset.filter(tlf_contacto='')

        if fb_ == '1':
            queryset = queryset.exclude(fanpage_fb='')
        elif fb_ == '2':
            queryset = queryset.filter(fanpage_fb='')

        if email_ == '1':
            queryset = queryset.exclude(emails='')
        elif email_ == '2':
            queryset = queryset.filter(emails='')

        if web_ == '1':
            queryset = queryset.exclude(website_posible='')
        elif web_ == '2':
            queryset = queryset.filter(website_posible='')

        data_ = list(queryset.values_list()) #8 por 11
        data_ = [(i[1],i[2],i[3],i[4],i[5],i[11],i[9], i[10],i[8],i[7],
                  i[12], i[23], i[13], i[14], i[15], i[16],i[17],i[18],i[19], i[20], i[21], i[22])
                for i in data_]
        number_items = len(data_)
        file_csv = save_data_csv(data_)
    return JsonResponse({'number_items':number_items, 'data': data_, 'file_csv': file_csv})
