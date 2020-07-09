#!/usr/bin/env python3
# import pandas as pd
import os, sys
sys.path.append('../../')
os.environ['DJANGO_SETTINGS_MODULE'] = 'networking.settings'
import django
django.setup()
from app.models import ModelNetworking
from openpyxl import *


def save_data(excel_file):
	net = pd.read_excel(os.path.join('data', excel_file))

	nc = net['RAZON SOCIAL']
	rp = net['REPRESENTANTES PRINCIPALES']
	po = net['PRESENCIA ONLINE']
	fa = net['FECHA DE APERTURA']
	te = net['TIPO DE EMPRESA']
	rtu = net['REPORT TECH URL']
	lcu = net['LINKEDIN CONTACT URL']
	ctg = net['CATEGORIA']
	tlf = net['TLF MOBILE']
	mai = net['EMAIL']
	web = net['WEBSITE ESTIMADO']
	fr = net['FUENTES RELACIONADAS']
	fp = net['FACEBOOK FANPAGE']
	ws = net['WEB SERVER']
	wf = net['WEB FRAMEWORK']
	jsf = net['JS FRAMEWORK']
	cms = net['CMS']
	emp = net['N° Empleados/Fecha']
	ruc = net['RUC']
	rub = net['RUBRO']
	dicc = net['Direccion']
	lp = net['LENGUAJE PROGRAMMING']
	stc = net['STATUS COMPANY']
	# twp = net['TWITTER PAGE']


	for data in zip(nc ,rp ,po ,fa, te ,rtu ,lcu ,ctg ,tlf ,mai ,web ,fr ,fp ,ws ,wf, jsf,cms ,emp ,ruc ,rub ,dicc ,lp ,stc ):
		data = list(data)
		for n in range(23):
			if str(data[n]) == 'nan' or str(data[n]) == 'Desconocido':
				data[n] = ''

		ModelNetworking.objects.create(
										name_company=data[0],
										representantes=data[1],
										presence_online=data[2],
										fecha_company=data[3],
										tamano_company=data[4],
										uri_reporte_tech=data[5],
										uri_reporte_linkedin=data[6],
										category=data[7],
										tlf_contacto=data[8],
										emails=data[9],
										website_posible=data[10],
										top_sites=data[11],
										fanpage_fb=data[12],
										web_server=data[13],
										web_framework=data[14],
										js_framework=data[15],
										cms=data[16],
										number_employers=data[17],
										ruc=data[18],
										rubro=data[19],
										direccion=data[20],
										programming_languaje=data[21],
										company_status=data[22]
		)
		print(data[18], ', created.')

# ModelNetworking.objects.all().delete()
#
# for i in range(4):
# 	try:
# 		fichero = 'networking_bot_{}.xlsx'.format(i)
# 		save_data(fichero)
# 	except Exception as e:
# 		print(e)


def dump_csv():
	queryset = ModelNetworking.objects.filter(presence_online__icontains = 'Alt')
	queryset2 = ModelNetworking.objects.filter(presence_online__icontains = 'Medi')

	wb = Workbook()
	ws = wb.active

	ws['A1'] = 'RAZON SOCIAL'
	ws['B1'] = 'REPRESENTANTES PRINCIPALES'
	ws['C1'] = 'PRESENCIA ONLINE'
	ws['D1'] = 'FECHA DE APERTURA'
	ws['E1'] = 'TIPO DE EMPRESA'
	ws['F1'] = 'REPORT TECH URL'
	ws['G1'] = 'LINKEDIN CONTACT URL'
	ws['H1'] = 'CATEGORIA'
	ws['I1'] = 'TLF MOBILE'
	ws['J1'] = 'EMAIL'
	ws['K1'] = 'WEBSITE ESTIMADO'
	ws['L1'] = 'FUENTES RELACIONADAS'
	ws['M1'] = 'FACEBOOK FANPAGE'
	ws['N1'] = 'WEB SERVER'
	ws['O1'] = 'WEB FRAMEWORK'
	ws['P1'] = 'JS FRAMEWORK'
	ws['Q1'] = 'CMS'
	ws['R1'] = 'N° Empleados/Fecha'
	ws['S1'] = 'RUC'
	ws['T1'] = 'RUBRO'
	ws['U1'] = 'Direccion'
	ws['V1'] = 'LENGUAJE PROGRAMMING'
	ws['W1'] = 'STATUS COMPANY'
	ws['X1'] = 'TWITTER PAGE'


	for dt in queryset:
		data = [dt.name_company ,
				dt.representantes ,
				dt.presence_online ,
				dt.fecha_company ,
				dt.tamano_company ,
				dt.uri_reporte_tech ,
				dt.uri_reporte_linkedin ,
				dt.category ,
				dt.tlf_contacto ,
				dt.emails ,
				dt.website_posible ,
				dt.top_sites ,
				dt.fanpage_fb ,
				dt.web_server ,
				dt.web_framework ,
				dt.js_framework ,
				dt.cms ,
				dt.number_employers ,
				dt.ruc ,
				dt.rubro ,
				dt.direccion ,
				dt.programming_languaje ,
				dt.company_status ,
				dt.twiter_page]
		try:
			ws.append(data)
		except Exception as e:
			debug(e)


	for dt in queryset2:
		data = [dt.name_company ,
				dt.representantes ,
				dt.presence_online ,
				dt.fecha_company ,
				dt.tamano_company ,
				dt.uri_reporte_tech ,
				dt.uri_reporte_linkedin ,
				dt.category ,
				dt.tlf_contacto ,
				dt.emails ,
				dt.website_posible ,
				dt.top_sites ,
				dt.fanpage_fb ,
				dt.web_server ,
				dt.web_framework ,
				dt.js_framework ,
				dt.cms ,
				dt.number_employers ,
				dt.ruc ,
				dt.rubro ,
				dt.direccion ,
				dt.programming_languaje ,
				dt.company_status ,
				dt.twiter_page]
		try:
			ws.append(data)
		except Exception as e:
			debug(e)


	wb.save('clientes.xlsx')
dump_csv()
