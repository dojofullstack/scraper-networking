#!/usr/bin/env python3

__title__ = "Service Networking"
__author__ = 'Henry Vasquez Conde'
__email__ = 'lifehack.py@gmail.com'
__description__ = """busca posibles clientes, contactos para ofrecerles servicios de ecommerce-ai,
                     obteniendo datos de varias fuentes abiertas"""
__version__ = '1.0b'

import os, sys
import re
import subprocess
# import cfscrape
import requests as req
from bs4 import BeautifulSoup as bs
from openpyxl import Workbook
from subprocess import Popen, PIPE
from random import randint
from time import sleep
from search_info.search_info import SearchSiteCompany
from detect_tech import detect_tech
import logging
from logzero import logger, loglevel, logfile
import argparse
os.environ['DJANGO_SETTINGS_MODULE'] = 'networking.settings'
import django
django.setup()
from app.models import ModelNetworking
from EmailHarvester import EmailHarvester

""" Setup Argument Parameters """
# parser = argparse.ArgumentParser(description='Discovery LinkedIn')
# parser.add_argument('-s', '--start', help='start paginate', required=1)
# parser.add_argument('-e', '--end', help='end paginate', required=1)
# args = parser.parse_args()
#
# start_page, end_page = int(args.start), int(args.end)
#
loglevel(logging.DEBUG) # debug in dev, warning in producction
debug = logger.debug
# warn = logger.warn
# logfile("/tmp/networking_bot.log", maxBytes=1e6, backupCount=3)

def info_contacto_website(web):
    if not('http' in web):
        if 'www' in web:
            web = 'http://{}'.format(web)
        elif not ('www' in web):
            web = 'http://www.{}'.format(web)

    # +56 2 2551 59 88 no detectado
    try:
        numeros, emails,facebook_pages, twitter_pages  = [], [], [], []
        page = req.get(web, timeout=60).content
        html = bs(page, 'html.parser')
        email_pattern = r'[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,4}'
        emails = re.findall(email_pattern, str(html))
        number_tlf = re.findall(r"\+\d{0,3}\s?0?\d{7,10}" , str(html))
        number_tlf2 = re.findall(r"\+?\d{0,3}?\s?0?\d{3}\s\d{3}\s\d{3}" , str(html))
        number_tlf3 =  re.findall(r"\+?\(?\d{0,3}\)?\s?0?\d{3}\s\d{4}", str(html))
        facebook_pages = re.findall(r"facebook.com[-\.A-Za-z0-9/]+", str(html))
        twitter_pages = re.findall(r"twitter.com[-\._A-Za-z0-9/]+", str(html))
        if number_tlf:
            numeros.extend(number_tlf)
        if number_tlf2:
            numeros.extend(number_tlf2)
        if number_tlf3:
            numeros.extend(number_tlf3)
        numeros = [n.replace(' ', '').strip() for n in numeros]
        numeros = list(set(numeros))
        facebook_pages = list(set(facebook_pages))
        twitter_pages = list(set(twitter_pages))
    except Exception as e:
        debug(e)
    try:
        more_emails = EmailHarvester.runing(web.replace('www.', '').replace('http://', '').replace('https://', '').replace('/', ''))
        emails.extend(more_emails)
    except Exception as e:
        debug(e)
    emails = list(set(emails))
    if emails:
        emails = [im for im in emails if not('22' in im)]
    print(numeros, emails, facebook_pages, twitter_pages)
    return(numeros, emails, facebook_pages, twitter_pages)


def upload_contactos_linkedin():
    cmd = 'curl --upload-file /home/henry/conctact_linkedin.xlsx https://transfer.sh/conctact_linkedin.xlsx'
    proc = Popen(cmd, shell=True, stdout=PIPE, stderr=PIPE, stdin=PIPE)
    uri_reporte = str(proc.stdout.read())
    return uri_reporte


# def search_univ_peru(ruc):
#     url = 'https://www.universidadperu.com/empresas/busqueda/'
#     data = {
#         'buscaempresa': ruc
#     }
#     try:
#         web_found, rubro_found = None, None
#         scraper = cfscrape.create_scraper()
#         page = scraper.post(url, data=data).content
#         html = bs(page, 'html.parser')
#         web = html.find('b', text=re.compile('Página Web:'))
#         web_found = web.find_next_sibling('a').text.strip()
#         rubro = html.find('b', text=re.compile('Actividad Comercial:'))
#         rubro_found = rubro.find_next_sibling('a').text.strip()
#         return(web_found, rubro_found)
#     except Exception as e:
#         # debug('error universidadperu',e)
#         return(web_found, rubro_found)


def list_to_str(mylist):
    if not mylist:
        return ''
    strin = ''
    for i in mylist:
        _ = '{}, \n '.format(i)
        strin += _
    return strin

def containt_facebook(top_sites):
    for fuente in top_sites:
        if 'facebook' in fuente:
            return fuente
    return ''

def clear_name(name):
    name = name.lower()
    filtros = ['e.i.r.l', 's.c.r.l', 's.r.ltda', 's.a.c', 's.r.l', 'H S.R.L']
    for filt in filtros:
        name = name.replace(filt, '')
    return name.strip()

def presence_online(top_sites, website, fanpage_fb):
    if website:
        return 'alto'
    elif fanpage_fb:
        return 'medio'
    elif len(top_sites) >= 3:
        return 'bajo'
    else:
        return ''

def get_owner_data(owner_url):
    if delay:
        time_delay = randint(1, delay_max)
        sleep(time_delay)
    try:
        debug('crawling: {}'.format(owner_url))
        url = 'https://datosperu.org/{}'.format(owner_url)
        page = req.get(url).content
        html = bs(page, 'html.parser')
        ruc = html.find('span', attrs={'itemprop': 'taxID'}).text.strip()
        name = html.find('span', attrs={'itemprop': 'name'}).text.strip()
        telef = html.find_all('span', attrs={'itemprop': 'telephone'})
        try:
            telef = [i.text for i in telef if len(i.text)>4]
        except:
            pass
        rubro = html.find('div', 'col-sm-12').text.strip()
        data_list = html.find('div', 'contact-details').text.strip()
        profile = html.find('div', 'company-profile').text.strip()
        profile = profile.split('\n')
        representantes_principales = []
        for i in profile:
            try:
                i = i.lower()
                if 'gerente general' in i:
                    gere = i.replace('gerente general', '').split('(desde')[0].strip()
                    representantes_principales.append(gere)
                if 'apoderado' in i:
                    repre = i.replace('apoderado', '').split('(desde')[0].strip()
                    representantes_principales.append(repre)
                if 'titular-gerente' in i:
                    titular = i.replace('titular-gerente', '').split('(desde')[0].strip()
                    representantes_principales.append(titular)
            except Exception as e:
                debug(e)

        state_company = ''
        it_website = None

        if 'ESTADO' in data_list:
            if 'ACTIVO' in data_list:
                state_company = 'Activo'

        tamano_company, number_employers, numeros, emails, fanpage_fb, uri_reporte_linkedin = '','', '', '', '', ''
        facebook_pages, twitter_pages = [], []
        all_tech = ''
        web_server, programming_languaje, web_framework, js_framework, cms, uri_reporte = '', '', '', '', '',''
        addr = html.find('span', attrs={'itemprop': 'streetAddress'}).text.strip()
        start_company = html.find('span', attrs={'itemprop': 'foundingDate'}).text.strip()
        try:
            size = html.find('table', class_='table-hover')
            if size:
                size = size.find('tbody')
                row_data = size.find('tr').find_all('td')
                for _,i in enumerate(row_data):
                    i = i.text
                    if _ == 0:
                        number_employers = 'fecha: {i}, '
                    if _ == 1:
                        number_employers += '#trabajadores: {i}, '
                    if _ == 3:
                        number_employers += '#prestadors de servicio: {i}'
                        i = int(i)
                        if i < 10:
                            tamano_company = 'Micro Empresa'
                        elif i < 49:
                            tamano_company = 'Pequeña Empresa'
                        elif i < 200:
                            tamano_company = 'Mediana Empresa'
                        else:
                            tamano_company = 'Grande Empresa'
        except Exception as e:
            debug(e)

        numeros = telef
        name = clear_name(name)
        obj = SearchSiteCompany()
        top_sites, website = obj.run(name)
        fanpage_fb = containt_facebook(top_sites)

        if website:
            debug(website)
            numeros, emails, facebook_pages, twitter_pages  = info_contacto_website(website)
            if fanpage_fb:
                facebook_pages.append(fanpage_fb)
            if numeros: numeros.extend([telef])
            web_server, programming_languaje, web_framework, js_framework, cms, all_tech  = detect_tech(website)
            #print(numeros, emails, all_tech)
        # uri_reporte_linkedin = run_module_linkedin(name)
        factor_presence_online = presence_online(top_sites, website,fanpage_fb)

        data_csv = {
                    'nombre': name,
                    'representantes_principales': list_to_str(representantes_principales),
                    'factor_presence_online':factor_presence_online,
                    'start_company': start_company,
                    'tamano_company': tamano_company,
                    'uri_reporte_tech': uri_reporte,
                    'uri_reporte_linkedin':uri_reporte_linkedin,
                    'category': 'MODA / VESTIMENTA',
                    'tlf': list_to_str(numeros),
                    'emails': list_to_str(emails),
                    'website_posible': website,
                    'top_sites': list_to_str(top_sites),
                    'fanpage_fb': list_to_str(facebook_pages),
                    'web_server': web_server,
                    'web_framework': web_framework,
                    'js_framework': js_framework,
                    'cms': cms,
                    'number_employers': number_employers,
                    'ruc': ruc,
                    'rubro': rubro,
                    'direccion': addr,
                    'programming_languaje': programming_languaje,
                    'state_company': state_company,
                    'twiter_page': list_to_str(twitter_pages),
                    'all_tech': all_tech
        }
        # debug(data_csv)
        save_in_db(data_csv)
        return(data_csv)
    except Exception as e:
        debug(e)
        return None


def crawl_rubro(page_init, page_end):
    links = []
    for pg in range(page_init, page_end+1):
        try:
            # url = 'https://www.datosperu.org/actividad-venta-minorista-productos-textiles-calzado-zapaterias-52322-pagina-{0}.php'
            #url = 'https://www.datosperu.org/actividad-fabricacion-de-prendas-de-vestir-ropa-vestimenta-18100-pagina-{0}.php'
            url = 'https://www.datosperu.org/actividad-venta-minorista-productos-farmacia-medicinas-y-articulos-tocador-52310-pagina-{0}.php'
            url = url.format(pg)
            page = req.get(url).content
            html = bs(page, 'html.parser')
            res_list = html.find_all('div', class_='single-product')
            links.extend([x.a['href'] for x in res_list])
        except:
            pass

    debug('se encontraon {} companys'.format(len(links)))
    return links


# def crawl_brands(pages, keyword):
#     links = []
#     for i in range(1, pages):
#         url = f'https://www.datosperu.org/buscador_empresas.php?buscar={keyword}&pg={i}'
#         page = req.get(url).content
#         html = bs(page, 'html.parser')
#         res_list = html.find_all('div', class_='single-product')
#         links.extend([x.a['href'] for x in res_list])
#     debug(f'se encontraon {len(links)} companys')
#     return links[:2]
    # return links

def run_module_linkedin(key_word):
    uri_reporte = ''
    cmd = "/usr/bin/python2 /home/henry/git/vex-scrapers-networking/LinkedinCrawler/ScrapedIn.py -u"
    cmd = cmd.split()
    cmd.append(key_word)
    process = subprocess.Popen(cmd , stdout=subprocess.PIPE)
    output, error = process.communicate()
    process.wait()

    if os.path.exists('/home/henry/conctact_linkedin.xlsx'):
        uri_reporte = upload_contactos_linkedin()
        os.remove('/home/henry/conctact_linkedin.xlsx')
        debug(uri_reporte)
    debug(uri_reporte)
    return uri_reporte


def run(page_init, page_end):
    rang_page = '{}_{}'.format(page_init, page_end)
    brands = crawl_rubro(page_init, page_end)
    data = list(map(get_owner_data, brands))
    # data = list(filter(lambda x: x != None, data))
    # wb = Workbook()
    # ws = wb.active
    #
    # ws['A1'] = 'RAZON SOCIAL'
    # ws['B1'] = 'REPRESENTANTES PRINCIPALES'
    # ws['C1'] = 'PRESENCIA ONLINE'
    # ws['D1'] = 'FECHA DE APERTURA'
    # ws['E1'] = 'TIPO DE EMPRESA'
    # ws['F1'] = 'REPORT TECH URL'
    # ws['G1'] = 'LINKEDIN CONTACT URL'
    # ws['H1'] = 'CATEGORIA'
    # ws['I1'] = 'TLF MOBILE'
    # ws['J1'] = 'EMAIL'
    # ws['K1'] = 'WEBSITE ESTIMADO'
    # ws['L1'] = 'FUENTES RELACIONADAS'
    # ws['M1'] = 'FACEBOOK FANPAGE'
    # ws['N1'] = 'WEB SERVER'
    # ws['O1'] = 'WEB FRAMEWORK'
    # ws['P1'] = 'JS FRAMEWORK'
    # ws['Q1'] = 'CMS'
    # ws['R1'] = 'N° Empleados/Fecha'
    # ws['S1'] = 'RUC'
    # ws['T1'] = 'RUBRO'
    # ws['U1'] = 'Direccion'
    # ws['V1'] = 'LENGUAJE PROGRAMMING'
    # ws['W1'] = 'STATUS COMPANY'
    # ws['X1'] = 'TWITTER PAGE'
    #
    # for item in data:
    #     try:
    #         ws.append(list(item.values()))
    #     except Exception as e:
    #         debug(e)
    # wb.save('analytic_client_{}.xlsx'.format(rang_page))

def save_in_db(data):
    try:
        name_company = data['nombre']
        top_sites = data['top_sites']
        representantes_principales = data['representantes_principales']
        factor_presence_online = data['factor_presence_online']
        start_company = data['start_company']
        tamano_company = data['tamano_company']
        category = data['category']
        tlf = data['tlf']
        emails = data['emails']
        website_posible = data['website_posible']
        fanpage_fb = data['fanpage_fb']
        web_server = data['web_server']
        web_framework = data['web_framework']
        js_framework = data['js_framework']
        cms = data['cms']
        number_employers = data['number_employers']
        ruc = data['ruc']
        rubro = data['rubro']
        direccion = data['direccion']
        programming_languaje = data['programming_languaje']
        state_company = data['state_company']
        twiter_page = data['twiter_page']
        all_tech = data['all_tech']

        ModelNetworking.objects.create(
                                        name_company = name_company,
                                        representantes = representantes_principales,
                                        presence_online = factor_presence_online,
                                        fecha_company = start_company,
                                        tamano_company = tamano_company,
                                        category = category,
                                        tlf_contacto = tlf,
                                        emails = emails,
                                        website_posible = website_posible,
                                        top_sites = top_sites,
                                        fanpage_fb = fanpage_fb,
                                        web_server = web_server,
                                        web_framework = web_framework,
                                        js_framework = js_framework,
                                        cms = cms,
                                        number_employers = number_employers,
                                        ruc = ruc,
                                        rubro = rubro,
                                        direccion = direccion,
                                        programming_languaje = programming_languaje,
                                        company_status = state_company,
                                        twiter_page = twiter_page,
                                        all_tech = all_tech,
                                        )
        print('save in db.', ruc)
    except Exception as e:
        print(e)


def add_data():
    data = ModelNetworking.objects.all()
    for dt in  data:
        try:
            name_company = dt.name_company
            website_posible = dt.website_posible
            all_tech = dt.all_tech
            ruc = dt.ruc
            if website_posible and not all_tech:
                debug(website_posible)
                web_server, programming_languaje, web_framework, js_framework, cms,all_tech  = detect_tech(website_posible)
                numeros, emails, facebook_pages, twitter_pages  = info_contacto_website(website_posible)
                obj = ModelNetworking.objects.get(ruc=ruc)
                obj.web_server = web_server
                obj.programming_languaje = programming_languaje
                obj.web_framework = web_framework
                obj.js_framework = js_framework
                obj.cms = cms
                obj.all_tech = all_tech
                obj.tlf_contacto = numeros
                obj.emails = emails
                obj.fanpage_fb = facebook_pages
                obj.twiter_page = twitter_pages
                obj.save()
                debug('saved report tech')
        except Exception as e:
            debug(e)

    # debug(len(data))

delay_max = 60
delay = True   # activate in producction

# add_data()
start_page = 350
end_page = 400
run(start_page, end_page)
# info_contacto_website('http://vexsoluciones.com')
# run_module_linkedin('vex soluciones')
# detect_tech('http://www.vexsoluciones.com')
